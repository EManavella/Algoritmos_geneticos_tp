import random
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook

# definiciones y parámetros
NUM_BITS       = 30
COEF           = 2**NUM_BITS - 1
TAM_POBLACION  = 10
PROB_CROSSOVER = 0.75
PROB_MUTACION  = 0.05
T              = 4
ELITE_SIZE     = 2

METODO = "ruleta" # "ruleta", "torneo" o "elitismo"
GENERACIONES = [20, 100, 200]
GENERACIONES_ELITISMO = [100]


def bits_a_dec(cromosoma):
    # convierte una lista de bits a su valor decimal
    return int("".join(str(b) for b in cromosoma), 2)

def cromosoma_aleatorio():
    # genera un cromosoma aleatorio de NUM_BITS bits
    return [random.randint(0, 1) for _ in range(NUM_BITS)]

def funcion_objetivo(cromosoma):
    # convierte el cromosoma a decimal y calcula su valor objetivo
    x = bits_a_dec(cromosoma)
    return (x / COEF) ** 2

def fitness(poblacion):
    # calcula el fitness de cada individuo y normaliza la lista para que sume 1
    fobj  = [funcion_objetivo(ind) for ind in poblacion]
    total = sum(fobj)
    return [f / total for f in fobj]


def ruleta(poblacion, lista_fitness):
    """"
    MÉTODO DE SELECCIÓN POR RULETA
     - cada individuo ocupa una cantidad de casillas proporcional a su fitness
    """
    casillas = []
    # cada individuo ocupa una cantidad de casillas proporcional a su fitness y no fuerza a que sean 100 posiciones
    for f in lista_fitness:
        cantidad = round(f * 100)
        # si queda en 0, le asignamos 1 para que tenga al menos una casilla y no quede fuera de la selección
        if cantidad == 0:
            cantidad = 1
        casillas.append(cantidad)
    array_ruleta = []
    # para cada indice y su cantidad de casillas, agregamos el indice a la ruleta esa cantidad de veces
    for indice, cantidad in enumerate(casillas):
        for _ in range(cantidad):
            array_ruleta.append(indice)
    # seleccionamos dos padres al azar de la ruleta
    padre1 = poblacion[array_ruleta[random.randint(0, len(array_ruleta) - 1)]]
    padre2 = poblacion[array_ruleta[random.randint(0, len(array_ruleta) - 1)]]
    return padre1, padre2



def torneo(poblacion, lista_fitness):
    """"
    MÉTODO DE SELECCIÓN POR TORNEO
        - se seleccionan T individuos al azar y el que tenga el mayor fitness es el ganador
    """
    def un_torneo(): 

        #selecciona T individuo s al azar 
        indices = random.sample(range(len(poblacion)), T)
        #max lo asignamos como el fitness del primer indiv que aparece en la lista de indices
        max = lista_fitness[indices[0]]
        ganador = indices[0]
        #loopeamos la lista a ver si hay un fitness mayor 
        for i in indices:
            if lista_fitness[i] > max:
                max = lista_fitness[i]
                ganador = i
        return poblacion[ganador]
    #retorna dos padres ganadores de dos torneos distintos
    return un_torneo(), un_torneo()

def seleccionar(poblacion, lista_fitness):
    # dependiendo del método elegido, se llama a la función de selección correspondiente
    if METODO == "ruleta":
        return ruleta(poblacion, lista_fitness)
    elif METODO == "torneo":
        return torneo(poblacion, lista_fitness)


def crossover1punto(padre1, padre2):
    # con probabilidad PROB_CROSSOVER, se elige un punto de cruce al azar y se intercambian las partes de los padres para crear dos hijos
    if random.random() <= PROB_CROSSOVER:
        punto = random.randint(1, NUM_BITS - 1)
        hijo1 = padre1[:punto] + padre2[punto:]
        hijo2 = padre2[:punto] + padre1[punto:]
        return hijo1, hijo2
    return padre1[:], padre2[:]

def mutacion(cromosoma):
    #copia la lista para no modificar el original
    cromosoma_mutado = cromosoma[:]

    if random.random() <= PROB_MUTACION:
        #el limite inferior no puede ser el ultimo bit porque sino no hay mutación
        limite_inferior = random.randint(0, NUM_BITS - 2)
        limite_superior = random.randint(limite_inferior + 1, NUM_BITS - 1)
        #el slincing no toma en cuenta el último indice, por eso se le suma 1
        # pisa lo que hay de i:j con lo mismo pero recorrido al revés 
        cromosoma_mutado[limite_inferior:limite_superior+1] = cromosoma_mutado[limite_inferior:limite_superior+1][::-1]
      
    return cromosoma_mutado


def estadisticas(poblacion, lista_fitness):
    # calcula el valor objetivo de cada individuo, luego obtiene el máximo, mínimo, promedio, desviación estándar y el cromosoma del máximo
    valores    = [funcion_objetivo(ind) for ind in poblacion]
    maximo     = max(valores)
    minimo     = min(valores)
    promedio   = sum(valores) / len(valores)
    promedio_f = sum(lista_fitness) / len(lista_fitness) #promedio de fitness que se usa para calcular la varianza
    varianza   = sum((f - promedio_f) ** 2 for f in lista_fitness) / len(lista_fitness) 
    desv_std   = varianza ** 0.5
    idx_max    = valores.index(maximo) # índice del individuo con el valor objetivo máximo, que se usará para obtener su cromosoma
    return {
        "maximo":    maximo,
        "minimo":    minimo,
        "promedio":  promedio,
        "desv_std":  desv_std,
        "cromosoma": poblacion[idx_max]
    }


def correr_ag(n_ciclos):
    #genera poblacion inicial
    poblacion     = [cromosoma_aleatorio() for _ in range(TAM_POBLACION)]
    #calcula el fitness de la poblacion inicial y guarda las estadisticas y el tiempo inicial
    lista_fitness = fitness(poblacion)
    historial     = [estadisticas(poblacion, lista_fitness)]
    t0 = time.perf_counter()
    # repite 20/100/200 veces el proceso de selección, crossover y mutación para generar nuevas poblaciones, calculando el fitness y guardando las estadísticas en cada ciclo
    for _ in range(n_ciclos - 1):
        nueva_poblacion = []
        while len(nueva_poblacion) < TAM_POBLACION:
            padre1, padre2 = seleccionar(poblacion, lista_fitness)
            hijo1, hijo2   = crossover1punto(padre1, padre2)
            hijo1          = mutacion(hijo1)
            hijo2          = mutacion(hijo2)
            nueva_poblacion.append(hijo1)
            if len(nueva_poblacion) < TAM_POBLACION:
                nueva_poblacion.append(hijo2)
        poblacion     = nueva_poblacion[:]
        lista_fitness = fitness(poblacion)
        historial.append(estadisticas(poblacion, lista_fitness))
    #me devuelve el historial de estadísticas y el tiempo total transcurrido desde t0 hasta el final del proceso
    return historial, time.perf_counter() - t0

def correr_ag_elitismo(n_ciclos):
    poblacion     = [cromosoma_aleatorio() for _ in range(TAM_POBLACION)]
    lista_fitness = fitness(poblacion)
    historial     = [estadisticas(poblacion, lista_fitness)]
    t0 = time.perf_counter()
    for _ in range(n_ciclos):
        # ordena la población por fitness de mayor a menor y selecciona los mejores individuos para pasarlos directamente a la nueva población
        pares_ordenados = sorted(zip(lista_fitness, poblacion), reverse=True)
        elite = [ind[:] for _, ind in pares_ordenados[:ELITE_SIZE]]
        nueva_poblacion = elite[:]
        # mientras la nueva población no alcance el tamaño deseado, se generan nuevos individuos mediante selección, crossover y mutación
        while len(nueva_poblacion) < TAM_POBLACION:
            padre1, padre2 = ruleta(poblacion, lista_fitness)
            hijo1, hijo2   = crossover1punto(padre1, padre2)
            hijo1          = mutacion(hijo1)
            hijo2          = mutacion(hijo2)
            nueva_poblacion.append(hijo1)
            if len(nueva_poblacion) < TAM_POBLACION:
                nueva_poblacion.append(hijo2)
        # actualiza la población con la nueva generación y calcula su fitness y estadísticas
        poblacion     = nueva_poblacion[:]
        lista_fitness = fitness(poblacion)
        historial.append(estadisticas(poblacion, lista_fitness))
    return historial, time.perf_counter() - t0


def graficar(historial, n_ciclos, metodo):
    # extrae los valores de máximo, mínimo y promedio de cada generacion para graficarlos
    ciclos    = list(range(len(historial)))
    maximos   = [h["maximo"]   for h in historial]
    minimos   = [h["minimo"]   for h in historial]
    promedios = [h["promedio"] for h in historial]

    plt.figure(figsize=(10, 5))
    plt.plot(ciclos, maximos,   label="Máximo",   color="steelblue", linewidth=2)
    plt.plot(ciclos, promedios, label="Promedio", color="seagreen",  linewidth=2)
    plt.plot(ciclos, minimos,   label="Mínimo",   color="tomato",    linewidth=2, linestyle="--")
    plt.title(f"AG {metodo.capitalize()} — {n_ciclos} generaciones")
    plt.xlabel("Generación")
    plt.ylabel("F. objetivo")
    plt.ylim(0, 1.05)
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    nombre = f"ag_{metodo}_{n_ciclos}_generaciones.png"
    plt.savefig(nombre, dpi=150)
    plt.show()
    print(f"  → Gráfica guardada: {nombre}")


def exportar_excel(todos_resultados, tiempos, metodo):
    #se exportan los resultados a excel
    wb = Workbook()
    wb.remove(wb.active)

    encabezados = ["Generacion", "Maximo", "Minimo", "Promedio", "Desv. Std", "Cromosoma del maximo"]

    for n_ciclos, historial in todos_resultados.items():
        ws = wb.create_sheet(title=f"{n_ciclos} generaciones")
        ws.append(encabezados)
        for i, stats in enumerate(historial):
            ws.append([
                i,
                round(stats["maximo"], 10),
                round(stats["minimo"], 10),
                round(stats["promedio"], 10),
                round(stats["desv_std"], 10),
                str(stats["cromosoma"]),
            ])

    ws_res = wb.create_sheet(title="Resumen tiempos")
    ws_res.append(["Generaciones", "Metodo", "Tiempo (s)"])
    for n, t in tiempos.items():
        ws_res.append([int(n), metodo.capitalize(), round(t, 6)])

    wb.move_sheet("Resumen tiempos", offset=-len(wb.sheetnames) + 1)

    nombre = f"resultados_{metodo}.xlsx"
    wb.save(nombre)



tiempos          = {}
todos_resultados = {}

gens = GENERACIONES_ELITISMO if METODO == "elitismo" else GENERACIONES

for n in gens:
    print(f"\n  Ejecutando {n} generaciones — Método: {METODO.capitalize()}...")
    if METODO == "elitismo":
        historial, tiempo = correr_ag_elitismo(n)
    else:
        historial, tiempo = correr_ag(n)
    tiempos[n]          = tiempo
    todos_resultados[n] = historial
    graficar(historial, n, METODO)
    print(f"  Tiempo de ejecución: {tiempo:.6f} s")

exportar_excel(todos_resultados, tiempos, METODO)
print(f"\n   Fin — Método: {METODO.capitalize()}")