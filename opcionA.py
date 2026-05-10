import random
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


NUM_BITS       = 30
COEF           = 2**NUM_BITS - 1
TAM_POBLACION  = 10
PROB_CROSSOVER = 0.75
PROB_MUTACION  = 0.05

GENERACIONES = [20, 100, 200]   

COLOR_HEADER    = "1F4E79"   # azul oscuro
COLOR_GEN0      = "D6E4F0"   # celeste claro para generación 0
COLOR_RESUMEN   = "1F4E79"   # azul oscuro para resumen
COLOR_FILA_PAR  = "EBF5FB"   # celeste muy claro filas pares
FUENTE          = "Arial"


def bits_a_dec(cromosoma):
    return int("".join(str(b) for b in cromosoma), 2)

def cromosoma_aleatorio():
    return [random.randint(0, 1) for _ in range(NUM_BITS)]

def funcion_objetivo(cromosoma):
    x = bits_a_dec(cromosoma)
    return (x / COEF) ** 2

def fitness(poblacion):
    fobj  = [funcion_objetivo(ind) for ind in poblacion]
    total = sum(fobj)
    return [f / total for f in fobj]


def ruleta(poblacion, lista_fitness):
    casillas = []
    for f in lista_fitness:
        cantidad = round(f * 100)
        if cantidad == 0:
            cantidad = 1
        casillas.append(cantidad)

    array_ruleta = []
    for indice, cantidad in enumerate(casillas):
        for _ in range(cantidad):
            array_ruleta.append(indice)

    padre1 = poblacion[array_ruleta[random.randint(0, len(array_ruleta) - 1)]]
    padre2 = poblacion[array_ruleta[random.randint(0, len(array_ruleta) - 1)]]
    return padre1, padre2

def crossover1punto(padre1, padre2):
    if random.random() <= PROB_CROSSOVER:
        punto = random.randint(1, NUM_BITS - 1)
        hijo1 = padre1[:punto] + padre2[punto:]
        hijo2 = padre2[:punto] + padre1[punto:]
        return hijo1, hijo2
    return padre1[:], padre2[:]

def mutacion(cromosoma):
    cromosoma = cromosoma[:]
    for i in range(len(cromosoma)):
        if random.random() <= PROB_MUTACION:
            cromosoma[i] = 1 - cromosoma[i]
    return cromosoma


def estadisticas(poblacion, lista_fitness):
    valores    = [funcion_objetivo(ind) for ind in poblacion]
    maximo     = max(valores)
    minimo     = min(valores)
    promedio   = sum(valores) / len(valores)
    promedio_f = sum(lista_fitness) / len(lista_fitness)
    varianza   = sum((f - promedio_f) ** 2 for f in lista_fitness) / len(lista_fitness)
    desv_std   = varianza ** 0.5
    idx_max    = valores.index(maximo)
    return {
        "maximo":    maximo,
        "minimo":    minimo,
        "promedio":  promedio,
        "desv_std":  desv_std,
        "cromosoma": poblacion[idx_max]
    }


def correr_ag(n_ciclos):
    poblacion     = [cromosoma_aleatorio() for _ in range(TAM_POBLACION)]
    lista_fitness = fitness(poblacion)
    historial     = [estadisticas(poblacion, lista_fitness)]   

    t0 = time.perf_counter()                  

    for _ in range(n_ciclos):
        nueva_poblacion = []
        while len(nueva_poblacion) < TAM_POBLACION:
            padre1, padre2 = ruleta(poblacion, lista_fitness)
            hijo1, hijo2   = crossover1punto(padre1, padre2)
            hijo1          = mutacion(hijo1)
            hijo2          = mutacion(hijo2)
            nueva_poblacion.append(hijo1)
            if len(nueva_poblacion) < TAM_POBLACION:
                nueva_poblacion.append(hijo2)

        poblacion     = nueva_poblacion[:]
        lista_fitness = fitness(poblacion)
        historial.append(estadisticas(poblacion, lista_fitness))

    tiempo = time.perf_counter() - t0

    return historial, tiempo


def imprimir_tabla(historial, n_ciclos):
    ancho = 64
    print(f"\n{'═' * ancho}")
    print(f"  TABLA — {n_ciclos} generaciones  |  Método: Ruleta")
    print(f"{'═' * ancho}")
    print(f"  {'Gen':>5} | {'Máximo':>10} | {'Mínimo':>10} | {'Promedio':>10} | {'Desv.Std':>10}")
    print(f"  {'─' * 58}")
    for i, stats in enumerate(historial):
        print(f"  {i:>5} | {stats['maximo']:>10.6f} | {stats['minimo']:>10.6f} | "
              f"{stats['promedio']:>10.6f} | {stats['desv_std']:>10.6f}")
    print(f"{'═' * ancho}")
    mejor = max(historial, key=lambda s: s["maximo"])
    print(f"  Mejor F.Obj : {mejor['maximo']:.6f}")
    print(f"  Cromosoma   : {mejor['cromosoma']}")
    print(f"  X           : {bits_a_dec(mejor['cromosoma'])}")


def graficar(historial, n_ciclos):
    ciclos    = list(range(len(historial)))
    maximos   = [h["maximo"]   for h in historial]
    minimos   = [h["minimo"]   for h in historial]
    promedios = [h["promedio"] for h in historial]

    plt.figure(figsize=(10, 5))
    plt.plot(ciclos, maximos,   label="Máximo",   color="steelblue", linewidth=2)
    plt.plot(ciclos, promedios, label="Promedio", color="seagreen",  linewidth=2)
    plt.plot(ciclos, minimos,   label="Mínimo",   color="tomato",    linewidth=2, linestyle="--")
    plt.title(f"AG Ruleta — {n_ciclos} generaciones")
    plt.xlabel("Generación")
    plt.ylabel("F. objetivo")
    plt.ylim(0, 1.05)
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    nombre = f"ag_ruleta_{n_ciclos}_generaciones.png"
    plt.savefig(nombre, dpi=150)
    plt.show()
    print(f"  → Gráfica guardada: {nombre}")

 
borde_fino = Border(
    left   = Side(style="thin", color="BFBFBF"),
    right  = Side(style="thin", color="BFBFBF"),
    top    = Side(style="thin", color="BFBFBF"),
    bottom = Side(style="thin", color="BFBFBF"),
)
 
def estilo_header(cell, color_fondo=COLOR_HEADER):
    cell.font      = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=color_fondo)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = borde_fino
 
def estilo_dato(cell, fila_par=False, negrita=False, color_fondo=None):
    fondo = color_fondo if color_fondo else (COLOR_FILA_PAR if fila_par else "FFFFFF")
    cell.font      = Font(name=FUENTE, bold=negrita, size=10)
    cell.fill      = PatternFill("solid", fgColor=fondo)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = borde_fino
    cell.number_format = "0.000000"
 
def ancho_columnas(ws, anchos):
    """anchos: dict { letra_col: ancho }"""
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
 
def exportar_excel(todos_resultados, tiempos):
    wb = Workbook()
    wb.remove(wb.active)  
 
    encabezados = ["Generación", "Máximo", "Mínimo", "Promedio", "Desv. Std"]
 
    # ── una hoja por cada cantidad de generaciones ──
    for n_ciclos, historial in todos_resultados.items():
        ws = wb.create_sheet(title=f"{n_ciclos} generaciones")
 
        # título de la hoja
        ws.merge_cells("A1:E1")
        titulo = ws["A1"]
        titulo.value     = f"AG Ruleta — {n_ciclos} generaciones"
        titulo.font      = Font(name=FUENTE, bold=True, size=13, color="FFFFFF")
        titulo.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
 
        # encabezados columnas
        for col, texto in enumerate(encabezados, start=1):
            cell = ws.cell(row=2, column=col, value=texto)
            estilo_header(cell)
        ws.row_dimensions[2].height = 20
 
        # datos fila por fila
        for i, stats in enumerate(historial):
            fila = i + 3   # fila 3 en adelante
            fila_par = (i % 2 == 0)
            color_fondo = COLOR_GEN0 if i == 0 else None   # gen 0 destacada
 
            valores_fila = [
                i,
                stats["maximo"],
                stats["minimo"],
                stats["promedio"],
                stats["desv_std"],
            ]
            for col, valor in enumerate(valores_fila, start=1):
                cell = ws.cell(row=fila, column=col, value=round(valor, 6) if col > 1 else int(valor))
                negrita = (i == 0)
                estilo_dato(cell, fila_par=fila_par, negrita=negrita, color_fondo=color_fondo)
                if col == 1:   
                    cell.number_format = "0"
 
        ancho_columnas(ws, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14})
 
    # ── hoja resumen de tiempos ──
    ws_res = wb.create_sheet(title="Resumen tiempos")
 
    ws_res.merge_cells("A1:C1")
    t_titulo = ws_res["A1"]
    t_titulo.value     = "Resumen de tiempos de cómputo — Ruleta"
    t_titulo.font      = Font(name=FUENTE, bold=True, size=13, color="FFFFFF")
    t_titulo.fill      = PatternFill("solid", fgColor=COLOR_RESUMEN)
    t_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 28
 
    for col, texto in enumerate(["Generaciones", "Método", "Tiempo (s)"], start=1):
        cell = ws_res.cell(row=2, column=col, value=texto)
        estilo_header(cell)
    ws_res.row_dimensions[2].height = 20
 
    for fila, (n, t) in enumerate(tiempos.items(), start=3):
        par = (fila % 2 == 0)
        for col, valor in enumerate([n, "Ruleta", round(t, 6)], start=1):
            cell = ws_res.cell(row=fila, column=col, value=valor)
            estilo_dato(cell, fila_par=par)
            if col == 1:
                cell.number_format = "0"
            if col == 3:
                cell.number_format = "0.000000"
 
    ancho_columnas(ws_res, {"A": 16, "B": 14, "C": 16})
 
    # mover hoja resumen al principio
    wb.move_sheet("Resumen tiempos", offset=-len(wb.sheetnames) + 1)
 
    nombre = "resultados_ruleta.xlsx"
    wb.save(nombre)
    print(f"\n  → Excel guardado: {nombre}")

tiempos = {}
todos_resultados = {}

for n in GENERACIONES:
    print(f"\n{'#'*64}")
    print(f"  AG con {n} generaciones — Método: Ruleta")
    print(f"{'#'*64}")

    historial, tiempo = correr_ag(n)
    tiempos[n] = tiempo
    todos_resultados[n] = historial

    imprimir_tabla(historial, n)
    graficar(historial, n)
    print(f"  Tiempo de ejecución: {tiempo:.6f} s")

# tabla final
print(f"\n{'═' * 42}")
print(f"  TABLA RESUMEN DE TIEMPOS")
print(f"{'═' * 42}")
print(f"  {'Generaciones':>12} | {'Método':>10} | {'Tiempo':>10}")
print(f"  {'─' * 38}")
for n in GENERACIONES:
    print(f"  {n:>12} | {'Ruleta':>10} | {tiempos[n]:>8.6f} s")
print(f"{'═' * 42}")


# exportar todo a Excel
exportar_excel(todos_resultados, tiempos)

print("\n  Fin de la opción A")